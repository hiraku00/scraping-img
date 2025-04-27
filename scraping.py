import requests
from bs4 import BeautifulSoup
import re
import json
from urllib.parse import urljoin, urlparse
import time
import logging
import argparse
from typing import Optional, Tuple, Any, Dict, List
import os
from io import BytesIO
from PIL import Image as PILImage
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import WebDriverException, TimeoutException
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# --- 設定 ---
HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36'}
DEFAULT_IMAGE_WIDTH = 100
DEBUG_LOG_FILE = "scraping_debug.log"
REQUEST_TIMEOUT = 20
SELENIUM_TIMEOUT = 30
POST_URL_SLEEP = 1.0

# --- Selenium Only ドメインリスト ---
SELENIUM_ONLY_DOMAINS = [
    "ebay.com",
    "mercari.com", # mercariはrequestsでも試行するが、Seleniumが必要な場合が多い
    "2ndstreet.jp", # 同上
]

# --- 固定する列ヘッダー名 ---
URL_HEADER_NAME = "URL"
IMAGE_URL_HEADER_NAME = "(work)画像URL"
IMAGE_EMBED_COL_LETTER = 'E' # 画像埋め込み列 (E列)
HYPHEN_COL_LETTER = 'D'      # ハイフン列 (D列)

# ============================================
# 1. WebDriver 管理 (共通基盤)
# ============================================
class WebDriverManager:
    """Selenium WebDriverの初期化と終了を管理するクラス"""
    def __init__(self):
        self.options = self._default_options()
        self.driver: Optional[webdriver.Chrome] = None

    def _default_options(self) -> Options:
        """WebDriverのデフォルトオプションを設定"""
        options = Options()
        logging.debug("WebDriver オプション初期化")
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--log-level=3") # WebDriverのログを抑制
        options.add_argument(f"user-agent={HEADERS['User-Agent']}")
        # Chrome自身のログを抑制 (効果は限定的)
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        return options

    def __enter__(self) -> Optional[webdriver.Chrome]:
        """WebDriverを初期化 (with文で使用)"""
        t_start = time.time()
        print("WebDriverを初期化しています (ヘッドレスモード)...")
        logging.info("WebDriverを初期化しています (ヘッドレスモード)...")
        try:
            # ここでChromeDriverのパスを指定する必要がある場合がある
            # 例: self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=self.options)
            # または webdriver.Chrome(executable_path='/path/to/chromedriver', options=self.options)
            self.driver = webdriver.Chrome(options=self.options)
            print("WebDriverの準備が完了しました。")
            logging.info(f"WebDriverの準備が完了しました。({time.time() - t_start:.3f}s)")
            return self.driver
        except WebDriverException as e:
            logging.error(f"WebDriverException: WebDriver準備失敗: {e}", exc_info=True)
            print(f"\nエラー: WebDriver準備失敗: {e}")
            if "cannot find chrome binary" in str(e).lower():
                print(">>> Chromeブラウザ本体が見つかりません。パスを確認してください。")
            elif "session not created" in str(e).lower() and "this version of chromedriver only supports chrome version" in str(e).lower():
                print(">>> ChromeDriverのバージョンとChrome本体のバージョンが一致していません。")
                print(">>> ChromeDriverを更新するか、対応するChromeバージョンを使用してください。")
            elif "executable needs to be in PATH" in str(e).lower():
                 print(">>> ChromeDriverの実行可能ファイルがPATH上にないか、指定されていません。")
            else:
                print(">>> ChromeDriverの準備またはChromeの起動に問題が発生しました。")
        except Exception as e:
            logging.error(f"予期せぬエラー: WebDriver準備中: {e}", exc_info=True)
            print(f"エラー: WebDriver準備中: {e}")
        return None # エラー時は None を返す

    def __exit__(self, exc_type, exc_val, exc_tb):
        """WebDriverを終了 (with文で使用)"""
        if self.driver:
            t_start = time.time()
            try:
                self.driver.quit()
                print("WebDriverを終了しました。")
                logging.info(f"WebDriverを終了しました。({time.time() - t_start:.3f}s)")
            except Exception as e:
                logging.error(f"WebDriver終了中のエラー: {e}", exc_info=True)
                print(f"エラー: WebDriver終了中: {e}")
        self.driver = None

# ============================================
# 2. HTML/JSON 解析ヘルパー (共通基盤)
# ============================================
def extract_meta_property(soup: BeautifulSoup, property_name: str) -> Optional[str]:
    """metaタグのproperty属性からcontentを取得"""
    tag = soup.find('meta', property=property_name)
    return tag['content'] if tag and tag.get('content') else None

def extract_meta_name(soup: BeautifulSoup, name: str) -> Optional[str]:
    """metaタグのname属性からcontentを取得"""
    tag = soup.find('meta', attrs={'name': name})
    return tag['content'] if tag and tag.get('content') else None

def find_image_in_json(json_obj: Any) -> Optional[str]:
    """JSONデータ構造から再帰的に画像URLを探す"""
    if not json_obj: return None
    image_url = None
    if isinstance(json_obj, dict):
        # Common patterns first
        if 'image' in json_obj:
            image_prop = json_obj['image']
            if isinstance(image_prop, str): image_url = image_prop
            elif isinstance(image_prop, list) and len(image_prop) > 0:
                first_item = image_prop[0]
                if isinstance(first_item, str): image_url = first_item
                elif isinstance(first_item, dict) and first_item.get('url'): image_url = first_item['url']
            elif isinstance(image_prop, dict) and image_prop.get('url'): image_url = image_prop['url']
        if image_url: return image_url

        # Check @graph (common in JSON-LD)
        if '@graph' in json_obj and isinstance(json_obj['@graph'], list):
            image_url = find_image_in_json(json_obj['@graph'])
            if image_url: return image_url

        # Recursively check other values
        for key, value in json_obj.items():
            # Avoid infinite loops for recursive structures if necessary
            if key != 'image' and key != '@graph': # Avoid re-checking keys already handled
                image_url = find_image_in_json(value)
                if image_url: return image_url

    elif isinstance(json_obj, list):
        for item in json_obj:
            image_url = find_image_in_json(item)
            if image_url: return image_url
    return None

def extract_json_ld_image(soup: BeautifulSoup) -> Optional[str]:
    """HTML内のJSON-LDスクリプトから画像URLを抽出"""
    scripts = soup.find_all('script', type='application/ld+json')
    for script in scripts:
        if script.string:
            try:
                json_data = json.loads(script.string)
                image_url = find_image_in_json(json_data)
                if image_url:
                    logging.info(f"Found JSON-LD image: {image_url[:60]}...")
                    return image_url
            except json.JSONDecodeError as e:
                logging.warning(f"JSON-LD parsing error: {e} - Content: {script.string[:100]}...")
            except Exception as e:
                logging.warning(f"Error processing JSON-LD: {e}")
    return None

def convert_to_absolute_path(base_url: str, target_path: str) -> str:
    """相対パスを絶対パスに変換"""
    if not target_path or target_path.startswith(('http://', 'https://', 'data:')):
        return target_path or ""
    if target_path.startswith('//'):
        scheme = urlparse(base_url).scheme
        return f"{scheme}:{target_path}"
    try:
        return urljoin(base_url, target_path)
    except Exception as e:
        logging.error(f"Absolute path conversion failed for base='{base_url}', target='{target_path}': {e}")
        return target_path # Fallback to original path

# ============================================
# 3. 画像処理 (共通基盤)
# ============================================
def download_and_prepare_image(image_url: str, target_width: int, referrer_url: Optional[str] = None) -> Optional[Tuple[BytesIO, int, int]]:
    """画像をダウンロードし、リサイズしてBytesIOオブジェクトで返す"""
    t_dl_start = time.time()
    try:
        logging.debug(f"Starting image download for: {image_url}")
        img_headers = HEADERS.copy()
        if referrer_url:
            img_headers['Referer'] = referrer_url
            logging.debug(f"Using Referer: {referrer_url}")

        # stream=True を使用し、大きな画像をメモリに一気に読み込まないようにする
        img_response = requests.get(image_url, headers=img_headers, stream=True, timeout=15)
        img_response.raise_for_status()
        logging.debug(f"Image download completed in {time.time() - t_dl_start:.3f}s. Status: {img_response.status_code}")

        content_type = img_response.headers.get('content-type')
        if not content_type or not content_type.lower().startswith('image/'):
            logging.warning(f"Non-image content type ({content_type}) for URL: {image_url}")
            return None

        # 画像データをBytesIOに読み込む
        img_data = BytesIO()
        for chunk in img_response.iter_content(chunk_size=8192):
            img_data.write(chunk)
        img_data.seek(0) # ポインタを先頭に戻す

        if img_data.getbuffer().nbytes == 0:
            logging.warning(f"Empty image data received for URL: {image_url}")
            return None

        t_proc_start = time.time()
        with PILImage.open(img_data) as img:
            # 画像モードの変換（必要に応じて）
            if img.mode == 'P': img = img.convert('RGBA')
            elif img.mode == 'CMYK': img = img.convert('RGB')
            elif img.mode == 'LA': img = img.convert('RGBA')

            original_width, original_height = img.size
            if original_width <= 0 or original_height <= 0:
                logging.warning(f"Invalid image dimensions ({original_width}x{original_height}) for URL: {image_url}")
                return None

            # リサイズ計算
            aspect_ratio = original_height / original_width
            target_height = max(1, int(target_width * aspect_ratio))

            logging.debug(f"Resizing image from {original_width}x{original_height} to {target_width}x{target_height}")
            img_resized = img.resize((target_width, target_height), PILImage.Resampling.LANCZOS)

            # 出力バッファとフォーマット決定
            output_buffer = BytesIO()
            # 元のフォーマットを尊重するが、WebPやGIFなどはPNGに変換
            save_format = img.format if img.format and img.format.upper() in ['JPEG', 'PNG', 'BMP', 'TIFF'] else 'PNG'
            if save_format == 'GIF': save_format = 'PNG' # ExcelはGIFを直接サポートしないことが多い
            # JPEGでRGBAモードの場合はRGBに変換
            if save_format == 'JPEG' and img_resized.mode in ['RGBA', 'LA', 'P']:
                logging.debug("Converting RGBA/LA/P image to RGB for JPEG saving.")
                img_resized = img_resized.convert('RGB')

            # 画像をバッファに保存
            img_resized.save(output_buffer, format=save_format, quality=85 if save_format == 'JPEG' else None)

        output_buffer.seek(0)
        if output_buffer.closed:
             logging.error(f"BytesIO buffer is closed unexpectedly after saving: {image_url}")
             return None

        logging.debug(f"Image processing completed in {time.time() - t_proc_start:.3f}s")
        return output_buffer, target_width, target_height

    except requests.exceptions.Timeout:
        logging.error(f"Image download timeout for URL: {image_url}")
        return None
    except requests.exceptions.RequestException as e:
        logging.error(f"Image download error for URL {image_url}: {e}", exc_info=False) # Keep log clean
        return None
    except PILImage.UnidentifiedImageError:
        logging.error(f"Cannot identify image file (PIL) for URL: {image_url}")
        return None
    except Exception as e:
        logging.error(f"Unexpected error during image processing for URL {image_url}: {e}", exc_info=True)
        return None

# ============================================
# 4. HTML 解析ロジック (ビジネスロジック + α)
# ============================================
# --- サイト固有の解析関数 ---
def _parse_okoku_image(soup: BeautifulSoup) -> Optional[str]:
    """買取王国 (okoku.jp) から画像URLを抽出"""
    product_image_div = soup.find('div', id='product_image')
    if product_image_div:
        slider_ul = product_image_div.find('ul', class_='bxslider')
        if slider_ul:
            first_li = slider_ul.find('li')
            if first_li:
                img_tag = first_li.find('img')
                if img_tag and img_tag.get('src'):
                    potential_url = img_tag['src']
                    if "noimage" not in potential_url.lower():
                        logging.info("Found image via okoku.jp specific logic (bxslider).")
                        return potential_url
                    else:
                        logging.debug("okoku.jp specific: Found img in product_image, but it seems to be a noimage placeholder.")
    return None

def _parse_2ndstreet_image(soup: BeautifulSoup) -> Optional[str]:
    """2nd Street (2ndstreet.jp) から画像URLを抽出"""
    goods_images_div = soup.find('div', id='goodsImages')
    if goods_images_div:
        img_tag = goods_images_div.find('img')
        if img_tag and img_tag.get('src'):
            potential_url = img_tag['src']
            # サムネイル (_mn.jpg や _tn.jpg) でないか確認
            if not potential_url.endswith(('_mn.jpg', '_tn.jpg')):
                logging.info("Found image via 2ndstreet.jp specific logic (goodsImages).")
                return potential_url
            else:
                logging.debug("2ndstreet.jp specific: Found img in goodsImages, but it seems to be a thumbnail.")
    return None

def _parse_mercari_image(soup: BeautifulSoup) -> Optional[str]:
    """Mercari (mercari.com) から画像URLを抽出"""
    # 1. __NEXT_DATA__ から試す (最も確実なことが多い)
    next_data_script = soup.find('script', id='__NEXT_DATA__', type='application/json')
    if next_data_script and next_data_script.string:
        try:
            next_data = json.loads(next_data_script.string)
            # データ構造は変更される可能性あり
            photos = next_data.get('props', {}).get('pageProps', {}).get('item', {}).get('photos', [])
            if photos and isinstance(photos, list) and len(photos) > 0 and isinstance(photos[0], str):
                image_url = photos[0]
                logging.info(f"Found image URL in __NEXT_DATA__ (Mercari): {image_url[:60]}...")
                # Mercariの場合、クエリパラメータが付いていることが多いが、そのまま利用
                return image_url
        except Exception as e:
            logging.warning(f"Error processing Mercari __NEXT_DATA__ JSON: {e}")

    # 2. フォールバック: 特徴的なalt属性を持つimgタグ
    mercari_img_alt = soup.find('img', alt=lambda x: x and 'のサムネイル' in x)
    if mercari_img_alt and mercari_img_alt.get('src') and 'static.mercdn.net' in mercari_img_alt['src']:
        img_src = mercari_img_alt['src']
        logging.info(f'Found specific img by alt (mercari - fallback): {img_src[:60]}...')
        return img_src # クエリパラメータを除去しない方が良い場合もある

    # 3. フォールバック: 特徴的なsrcパターンを持つimgタグ
    mercari_pattern = re.compile(r'https://static\.mercdn\.net/item/detail/orig/photos/[^"\']+?')
    img_tag_src = soup.find('img', src=mercari_pattern)
    if img_tag_src and img_tag_src.get('src'):
        img_src = img_tag_src['src']
        logging.info(f"Found specific img src (mercari pattern - fallback): {img_src[:60]}...")
        return img_src # クエリパラメータを除去しない方が良い場合もある

    return None

def _parse_amazon_image(soup: BeautifulSoup) -> Optional[str]:
    """Amazonから画像URLを抽出"""
    # プライマリイメージコンテナIDのリスト (優先度順)
    image_container_ids = ['imgTagWrapperId', 'landingImage', 'ivLargeImage', 'main-image-container']
    main_image_container = None
    for container_id in image_container_ids:
        main_image_container = soup.find(id=container_id)
        if main_image_container:
            logging.debug(f"Amazon specific: Found container with id '{container_id}'")
            break

    if main_image_container:
        main_img = main_image_container.find('img')
        if main_img:
            potential_src = main_img.get('src') or main_img.get('data-src') # data-srcも考慮
            if potential_src and not potential_src.startswith("data:image") and "captcha" not in potential_src.lower():
                logging.info(f"Found potential main image via ID/Tag (Amazon): {potential_src[:60]}...")
                # Amazonの画像URLは複雑なことが多く、?以降を除去すると表示されない場合もあるので注意
                # 一旦そのまま返すか、パターンを見て除去するか判断
                # 例: ._AC_SL1500_.jpg のようなサイズ指定部分を除去する試み
                cleaned_url = re.sub(r'\._[A-Z]{2}_\w+_\.', '.', potential_src)
                if cleaned_url != potential_src:
                     logging.debug(f"Cleaned Amazon URL: {cleaned_url[:60]}...")
                     return cleaned_url
                return potential_src.split("?")[0] # シンプルな除去
            else:
                 logging.debug(f"Amazon specific: Img tag found in container, but src is invalid or missing: {potential_src}")
        else:
             logging.debug(f"Amazon specific: No img tag found within container id '{container_id}'")

    return None

# --- 標準的な解析関数 ---
def _extract_standard_metadata_image(soup: BeautifulSoup, domain: str) -> Optional[str]:
    """og:image, twitter:image から画像URLを抽出"""
    image_url = extract_meta_property(soup, 'og:image')
    # 買取王国の場合、og:imageがロゴ画像のことがあるため除外
    if image_url and "og_logo.png" in image_url and "okoku.jp" in domain:
        logging.debug(f"Skipping og:image because it seems to be a logo (okoku.jp): {image_url}")
        image_url = None
    elif image_url:
        logging.info(f"Found og:image: {image_url[:60]}...")
        return image_url

    image_url = extract_meta_name(soup, 'twitter:image')
    if image_url:
        logging.info(f"Found twitter:image: {image_url[:60]}...")
        return image_url

    return None

def _extract_fallback_image(soup: BeautifulSoup) -> Optional[str]:
    """最終手段として、一般的なimgタグから画像を探す"""
    logging.debug("Applying generic img tag fallback logic.")
    checked_sources = set()
    exclude_patterns = [
        ".gif", ".svg", "ads", "icon", "logo", "sprite", "avatar", "spinner",
        "loading", "pixel", "fls-fe.amazon", "transparent", "spacer", "dummy",
        "captcha", "_tn.", "_mn.", "_thumb.", "/thumb", "-small.", ".small",
        "nav_", "banner", "profile", "badge", "button", "rating"
    ]
    exclude_extensions = ['.php', '.aspx', '.jsp'] # スクリプトを示す拡張子は除外

    candidate_images = []

    for img in soup.find_all('img'):
        potential_src = img.get('src') or img.get('data-src') # 遅延読み込みも考慮
        if not potential_src or potential_src in checked_sources:
            continue

        checked_sources.add(potential_src)
        src_lower = potential_src.lower()

        # 明らかな除外対象かチェック
        if (src_lower.startswith("data:image") or
                any(pat in src_lower for pat in exclude_patterns) or
                any(src_lower.endswith(ext) for ext in exclude_extensions) or
                len(potential_src) < 15): # 短すぎるURLは除外
            continue

        # サイズ情報を取得できれば評価に加える
        width = img.get('width')
        height = img.get('height')
        size_score = 0
        try:
            if width and height and width.isdigit() and height.isdigit():
                w, h = int(width), int(height)
                if w > 50 and h > 50: # 小さすぎる画像は除外
                   size_score = w * h
        except:
            pass # サイズ取得失敗は無視

        # alt属性もヒントにする (商品名などが入っている可能性)
        alt_text = img.get('alt', '').lower()
        alt_score = 1 if alt_text and 'thumbnail' not in alt_text and 'logo' not in alt_text else 0

        # 候補リストに追加 (URL, サイズスコア, altスコア)
        candidate_images.append((potential_src, size_score, alt_score))
        logging.debug(f"Fallback candidate: {potential_src[:60]}... (Size: {size_score}, Alt: {alt_score})")

    if not candidate_images:
        logging.debug("No suitable fallback image candidates found.")
        return None

    # 候補を評価して最適なものを選択 (例: サイズが大きいものを優先)
    candidate_images.sort(key=lambda x: (x[1], x[2]), reverse=True) # サイズスコア、altスコアで降順ソート
    best_image_url = candidate_images[0][0]

    logging.info(f"Selected fallback image: {best_image_url[:60]}...")
    # URLからクエリパラメータを除去 (場合による)
    return best_image_url.split("?")[0]


# --- HTML解析メイン関数 ---
def parse_html_for_image(html_content: str, base_url: str) -> Optional[str]:
    """HTMLコンテンツを解析して最適な画像URLを返す"""
    if not html_content:
        logging.warning("parse_html_for_image received empty HTML content.")
        return None

    t_start = time.time()
    logging.debug(f"Start parsing HTML for: {base_url}")
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
    except Exception as e:
        logging.error(f"BeautifulSoup parsing failed for {base_url}: {e}", exc_info=True)
        return None
    logging.debug(f"BeautifulSoup parsed in {time.time() - t_start:.3f}s")

    domain = urlparse(base_url).netloc.lower()
    image_url: Optional[str] = None

    # 1. サイト固有のロジック (優先度 高)
    logging.debug(f"Checking site-specific parsers for domain: {domain}")
    if "okoku.jp" in domain: image_url = _parse_okoku_image(soup)
    elif "2ndstreet.jp" in domain: image_url = _parse_2ndstreet_image(soup)
    elif "mercari.com" in domain: image_url = _parse_mercari_image(soup)
    elif "amazon." in domain: image_url = _parse_amazon_image(soup)
    # 他のサイト固有ロジックがあればここに追加
    # elif "example.com" in domain: image_url = _parse_example_image(soup)

    # 2. 標準的なメタデータ (og:image, twitter:image)
    if not image_url:
        logging.debug("Trying standard metadata (og:image, twitter:image)")
        image_url = _extract_standard_metadata_image(soup, domain)

    # 3. JSON-LD
    if not image_url:
        logging.debug("Trying JSON-LD extraction")
        image_url = extract_json_ld_image(soup)

    # 4. フォールバック (一般的な<img>タグ)
    if not image_url:
        logging.debug("Trying fallback image extraction from <img> tags")
        image_url = _extract_fallback_image(soup)

    # 最終的なURLの絶対パス変換と返却
    if image_url:
        final_url = convert_to_absolute_path(base_url, image_url)
        logging.info(f"Found image URL: {final_url}")
        logging.debug(f"parse_html completed in {time.time() - t_start:.3f}s")
        return final_url
    else:
        logging.warning(f"No suitable image URL found after all parsing attempts for: {base_url}")
        logging.debug(f"parse_html completed in {time.time() - t_start:.3f}s")
        return None

# ============================================
# 5. コア Web スクレイピング処理
# ============================================
def _get_image_url_with_selenium(driver: webdriver.Chrome, url: str) -> Optional[str]:
    """指定されたURLをSeleniumで開き、画像URLを抽出する"""
    image_url = None
    try:
        logging.info(f"Attempting to fetch URL with Selenium: {url}")
        t_get = time.time()
        # ページの読み込みタイムアウトを設定
        driver.set_page_load_timeout(SELENIUM_TIMEOUT)
        driver.get(url)
        logging.debug(f"Selenium driver.get() completed in {time.time() - t_get:.3f}s")

        # JavaScriptの実行や動的コンテンツの読み込みを待機 (必要に応じて調整)
        # time.sleep(2) # 固定待機 (あまり推奨されない)
        # WebDriverWait(driver, 10).until(...) # 特定の要素が現れるまで待機 (より堅牢)
        # ここでは短い固定待機を入れておく
        t_sleep = time.time(); wait_seconds = 0.5; logging.debug(f"Waiting {wait_seconds}s for dynamic content...")
        time.sleep(wait_seconds); logging.debug(f"Selenium wait completed in {time.time() - t_sleep:.3f}s")

        t_source = time.time()
        page_source = driver.page_source
        current_url = driver.current_url # リダイレクト後のURLを取得
        logging.debug(f"Selenium driver.page_source retrieved in {time.time() - t_source:.3f}s")
        logging.info(f"Successfully fetched page source with Selenium (Final URL: {current_url})")

        # 取得したHTMLソースを解析
        image_url = parse_html_for_image(page_source, current_url)

        if image_url:
            logging.info(f"Found image URL using Selenium for {url}")
        else:
            logging.warning(f"Could not find image URL even with Selenium for {url}")

    except TimeoutException:
        logging.error(f"Selenium page load timed out ({SELENIUM_TIMEOUT}s) for URL: {url}")
    except WebDriverException as e:
        logging.error(f"Selenium WebDriver error for URL {url}: {e}", exc_info=True)
        # ここで特定のWebDriverエラーに対するハンドリングを追加可能
        # 例: if "net::ERR_CONNECTION_REFUSED" in str(e): ...
    except Exception as e:
        logging.error(f"Unexpected error during Selenium processing for URL {url}: {e}", exc_info=True)

    return image_url

def get_image_url_from_url(url: str, driver: Optional[webdriver.Chrome] = None) -> Tuple[Optional[str], Optional[str]]:
    """
    指定されたURLから画像URLを取得するメイン関数。
    まずrequestsで試行し、失敗した場合や特定のドメインの場合はSeleniumを使用する。
    Returns:
        Tuple[Optional[str], Optional[str]]: (画像URL, エラーメッセージ)
    """
    final_image_url: Optional[str] = None
    error_message: Optional[str] = None
    t_start_url = time.time()

    parsed_url = urlparse(url)
    domain = parsed_url.netloc.lower()

    # --- Seleniumを直接使用するかどうかの判定 ---
    # SELENIUM_ONLY_DOMAINS に含まれるドメインは、最初からSeleniumを使う
    use_selenium_directly = any(d in domain for d in SELENIUM_ONLY_DOMAINS if d) # 空文字を除外

    if use_selenium_directly:
        logging.info(f"Domain '{domain}' requires Selenium. Using Selenium directly for {url}")
        if driver:
            final_image_url = _get_image_url_with_selenium(driver, url)
            if not final_image_url:
                error_message = "画像が見つかりません(Sel-Direct)"
        else:
            logging.warning(f"Selenium is required for {url}, but Selenium driver is not available.")
            error_message = "画像が見つかりません(NoDriver)"
        log_func = logging.info if final_image_url else logging.warning
        log_func(f"URL processing (Selenium Direct) took {time.time() - t_start_url:.3f}s for {url}")
        return final_image_url, error_message

    # --- requests での試行 ---
    logging.info(f"Attempting to fetch URL with requests: {url}")
    t_req_start = time.time()
    response: Optional[requests.Response] = None
    html_content: str = ""
    base_url: str = url

    try:
        response = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        logging.debug(f"requests.get() completed in {time.time() - t_req_start:.3f}s")

        # ステータスコードチェック
        response.raise_for_status()

        # エンコーディング設定 (文字化け対策)
        response.encoding = response.apparent_encoding
        html_content = response.text
        base_url = response.url # リダイレクト後のURLを使用
        logging.info(f"Requests success for {url} (Final: {base_url}) - Status: {response.status_code}")

        # HTML解析
        final_image_url = parse_html_for_image(html_content, base_url)

        if not final_image_url:
            error_message = "画像が見つかりません(Req)"
            logging.warning(f"Image not found with requests for: {url}")

    except requests.exceptions.Timeout:
        error_message = f"タイムアウト(>{REQUEST_TIMEOUT}s)(Req)"
        logging.warning(f"Requests Timeout (>{REQUEST_TIMEOUT}s) in {time.time() - t_req_start:.3f}s for {url}.")
    except requests.exceptions.RequestException as e:
        status_code = e.response.status_code if e.response is not None else "N/A"
        error_message = f"アクセス失敗(Code:{status_code})(Req)"
        logging.error(f"Requests Access Error ({time.time() - t_req_start:.3f}s) {url}: Status {status_code}, Error {e}")
    except Exception as e:
        # BeautifulSoupのパースエラーなどもここでキャッチされる可能性
        error_message = f"エラー(Req): {str(e)[:50]}"
        logging.error(f"Requests URL Processing Error ({time.time() - t_req_start:.3f}s) {url}: {e}", exc_info=True)
    finally:
        # レスポンスオブジェクトを閉じる (stream=True を使った場合などに有効)
        if response:
            response.close()

    # --- Seleniumでのリトライ ---
    # requestsで画像が見つからなかった、またはrequests自体が失敗した場合で、
    # かつSeleniumドライバが利用可能な場合にリトライする
    if not final_image_url and driver:
        logging.info(f"Requests failed or couldn't find image for {url}. Retrying with Selenium...")
        t_sel_start = time.time()
        selenium_image_url = _get_image_url_with_selenium(driver, url)
        logging.debug(f"_get_image_url_with_selenium (Retry) completed in {time.time() - t_sel_start:.3f}s")

        if selenium_image_url:
            final_image_url = selenium_image_url
            error_message = None # Seleniumで成功したのでエラーメッセージをクリア
        else:
            # Seleniumでも見つからなかった場合、エラーメッセージを更新または追記
            current_error = error_message if error_message else "取得エラー"
            error_message = f"{current_error} / 画像が見つかりません(Sel-Retry)"

    elif not final_image_url and not driver:
        logging.warning(f"Requests failed for {url}, and Selenium retry is unavailable (no driver).")
        if not error_message: # requestsは成功したが画像が見つからなかった場合
             error_message = "画像が見つかりません(Req, No Retry)"


    log_func = logging.info if final_image_url else logging.warning
    log_func(f"URL processing (Hybrid) took {time.time() - t_start_url:.3f}s for {url}. Result: {'Found' if final_image_url else 'Not Found'}")
    return final_image_url, error_message

# ============================================
# 6. Excel 処理 (ファイル入出力)
# ============================================
def load_workbook_and_sheet(file_path: str, sheet_identifier: Any) -> Tuple[Optional[Workbook], Optional[Worksheet]]:
    """Excelファイルを読み込み、指定されたシートを取得する"""
    if not os.path.isfile(file_path):
        print(f"エラー: ファイルが見つかりません: '{file_path}'")
        logging.error(f"Input file not found: {file_path}")
        return None, None
    if not file_path.lower().endswith('.xlsx'):
        print("エラー: 入力ファイルは .xlsx 形式である必要があります。")
        logging.error(f"Invalid file format: {file_path}. Must be .xlsx")
        return None, None

    t_excel_load = time.time()
    print(f"入力ファイル '{file_path}' を読み込んでいます...")
    try:
        workbook = openpyxl.load_workbook(file_path)
        logging.info(f"Excel file loaded in {time.time() - t_excel_load:.3f}s")

        sheet_name: Optional[str] = None
        if isinstance(sheet_identifier, int):
            if 0 <= sheet_identifier < len(workbook.sheetnames):
                sheet_name = workbook.sheetnames[sheet_identifier]
            else:
                raise ValueError(f"シートインデックス '{sheet_identifier}' が範囲外です。利用可能なシート数: {len(workbook.sheetnames)}")
        elif isinstance(sheet_identifier, str):
            if sheet_identifier in workbook.sheetnames:
                sheet_name = sheet_identifier
            else:
                raise ValueError(f"シート名 '{sheet_identifier}' が見つかりません。利用可能なシート: {workbook.sheetnames}")
        else:
             raise ValueError(f"無効なシート識別子: {sheet_identifier} (文字列または整数を指定してください)")

        sheet = workbook[sheet_name]
        print(f"処理対象シート: '{sheet.title}'")
        return workbook, sheet
    except ValueError as ve:
        print(f"シート選択エラー: {ve}")
        logging.error(f"Sheet selection error: {ve}")
        if 'workbook' in locals() and workbook: workbook.close() # 念のため閉じる
        return None, None
    except Exception as load_e:
        print(f"\nエラー: Excelファイルの読み込みに失敗しました: {load_e}")
        logging.error(f"Failed to load Excel file '{file_path}': {load_e}", exc_info=True)
        return None, None

def get_column_indices(sheet: Worksheet, required_headers: List[str]) -> Optional[Dict[str, int]]:
    """ヘッダー行を解析し、必要な列のインデックスを取得する"""
    header_row_index = 1 # ヘッダーは1行目にあると仮定
    headers: Dict[str, int] = {}
    try:
        for cell in sheet[header_row_index]:
            if cell.value is not None:
                headers[str(cell.value)] = cell.column
        logging.debug(f"Found headers: {headers}")

        # 必要なヘッダーが存在するかチェック
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            print(f"エラー: 必須の列ヘッダーが見つかりません: {missing_headers}")
            print(f"シート '{sheet.title}' の1行目にあるヘッダー: {list(headers.keys())}")
            logging.error(f"Missing required headers: {missing_headers}. Found headers: {list(headers.keys())}")
            return None

        return {h: headers[h] for h in required_headers}

    except Exception as e:
        print(f"エラー: ヘッダー行の解析中にエラーが発生しました: {e}")
        logging.error(f"Error parsing header row in sheet '{sheet.title}': {e}", exc_info=True)
        return None

def clear_previous_results(sheet: Worksheet, url_col_idx: int, img_url_col_idx: int, img_embed_col_letter: str, hyphen_col_letter: str):
    """指定された列の既存の処理結果をクリアする"""
    logging.info("Clearing previous results (Image URL, Embedded Image, Hyphen)...")
    img_embed_col_idx = openpyxl.utils.column_index_from_string(img_embed_col_letter)
    hyphen_col_idx = openpyxl.utils.column_index_from_string(hyphen_col_letter)

    for row_idx in range(2, sheet.max_row + 1): # ヘッダー行(1)を除く
        # 画像URL列
        img_url_cell = sheet.cell(row=row_idx, column=img_url_col_idx)
        img_url_cell.value = None
        img_url_cell.hyperlink = None
        img_url_cell.style = 'Normal' # スタイルをリセット

        # ハイフン列
        hyphen_cell = sheet.cell(row=row_idx, column=hyphen_col_idx)
        hyphen_cell.value = None

        # 画像埋め込み列 (値は通常Noneだが、念のため)
        img_embed_cell = sheet.cell(row=row_idx, column=img_embed_col_idx)
        img_embed_cell.value = None # 画像自体は後で削除

    # 既存の画像をシートから削除
    if sheet._images:
         print("注意: シート上の既存の画像をすべて削除します...")
         logging.warning("Removing all existing images from the sheet.")
         sheet._images = []
    logging.info("Previous results cleared.")


def process_excel_rows(
    sheet: Worksheet,
    url_col_idx: int,
    img_url_col_idx: int,
    img_embed_col_letter: str,
    hyphen_col_letter: str,
    image_width: int,
    sleep_interval: float,
    process_all_rows: bool,
    driver: Optional[webdriver.Chrome]
) -> int:
    """Excelの各行を処理し、画像を取得・埋め込みを行う"""
    processed_count = 0
    start_row = 2 # ヘッダー行の次から開始
    img_embed_col_idx = openpyxl.utils.column_index_from_string(img_embed_col_letter)
    hyphen_col_idx = openpyxl.utils.column_index_from_string(hyphen_col_letter)

    total_rows_with_urls = sum(1 for row_idx in range(start_row, sheet.max_row + 1)
                               if sheet.cell(row=row_idx, column=url_col_idx).value)
    print(f"処理対象のURLを含む行数: {total_rows_with_urls}")
    if total_rows_with_urls == 0:
        print("処理対象のURLが見つかりませんでした。")
        return 0

    overall_start_time = time.time() # ループ開始時間

    for row_index in range(start_row, sheet.max_row + 1):
        url_cell = sheet.cell(row=row_index, column=url_col_idx)
        url = str(url_cell.value).strip() if url_cell.value is not None else ""

        # URLがない場合
        if not url:
            if not process_all_rows:
                logging.debug(f"Row {row_index}: Skipping empty URL.")
                continue # process_all=False ならスキップ
            else:
                print(f"\n行 {row_index}: URLが空のため処理を中断します。")
                logging.info(f"Row {row_index}: Stopping processing due to empty URL (process_all=True).")
                break # process_all=True なら中断

        # URL形式チェック
        if not url.lower().startswith(('http://', 'https://')):
            logging.warning(f"Row {row_index}: Invalid URL format: {url}")
            error_msg = "無効なURL形式"
            sheet.cell(row=row_index, column=img_url_col_idx).value = error_msg
            # 無効URLの場合もハイフンを入れるか？ -> 入れない仕様とする
            # sheet.cell(row=row_index, column=hyphen_col_idx).value = "-" # 必要ならコメント解除
            if process_all_rows:
                print(f"\n行 {row_index}: 無効なURL形式のため処理を中断します。")
                logging.info(f"Row {row_index}: Stopping processing due to invalid URL (process_all=True).")
                break
            else:
                 continue # 次の行へ

        # --- ここから有効なURLの処理 ---
        processed_count += 1
        print(f"\r処理中: {processed_count}/{total_rows_with_urls} 件目 ({row_index}行目) - {url[:60]}...", end="", flush=True)
        logging.info(f"--- Processing Row {row_index}, URL: {url} ---")

        # D列にハイフンを設定
        sheet.cell(row=row_index, column=hyphen_col_idx).value = "-"

        # 画像URLを取得
        image_url, error_message = get_image_url_from_url(url, driver)

        img_url_cell = sheet.cell(row=row_index, column=img_url_col_idx)
        img_embed_cell = sheet.cell(row=row_index, column=img_embed_col_idx)

        if image_url:
            img_url_cell.value = image_url
            # ハイパーリンクを設定（Excelの制限に注意）
            try:
                 img_url_cell.hyperlink = image_url
                 img_url_cell.style = "Hyperlink"
            except Exception as e:
                 logging.warning(f"Row {row_index}: Failed to set hyperlink for image URL: {e}")

            # 画像をダウンロードして準備
            image_result = download_and_prepare_image(image_url, image_width, referrer_url=url)

            if image_result:
                image_data_buffer, img_width, img_height = image_result
                try:
                    if not image_data_buffer.closed:
                        # --- 画像埋め込み ---
                        img_for_excel = OpenpyxlImage(image_data_buffer)
                        img_for_excel.width = img_width
                        img_for_excel.height = img_height

                        # 行の高さ調整
                        required_row_height = img_height * 0.75 + 2 # ポイント単位に変換 + 余白
                        current_height = sheet.row_dimensions[row_index].height
                        if current_height is None or current_height < required_row_height:
                            sheet.row_dimensions[row_index].height = required_row_height
                            logging.debug(f"Row {row_index}: Set row height to {required_row_height:.2f}")

                        # 列の幅調整 (初回のみ or 必要なら毎回)
                        col_letter = img_embed_col_letter
                        required_col_width = img_width / 7.0 + 1 # Excelの幅単位に変換 + 余白 (調整が必要かも)
                        current_width = sheet.column_dimensions[col_letter].width
                        # 常に上書きする or 未設定か小さい場合のみ設定するか選択
                        # if current_width is None or current_width < required_col_width:
                        if current_width is None or current_width < required_col_width + 5: # 少し余裕を持たせる
                            sheet.column_dimensions[col_letter].width = required_col_width + 5
                            logging.debug(f"Row {row_index}: Set column {col_letter} width to {required_col_width + 5:.2f}")


                        # セルのアンカーと配置
                        cell_anchor = f"{img_embed_col_letter}{row_index}"
                        # セルの内容配置を中央揃えに (画像自体のアラインメントではない)
                        img_embed_cell.alignment = Alignment(horizontal='center', vertical='center')

                        sheet.add_image(img_for_excel, cell_anchor)
                        logging.info(f"Row {row_index}: Image successfully embedded into cell {cell_anchor}")
                    else:
                        logging.error(f"Row {row_index}: Image data buffer was closed before embedding.")
                        img_embed_cell.value = "内部エラー(Buffer)"
                except ValueError as ve:
                    # openpyxl がサポートしていない画像形式などの場合
                    logging.error(f"Row {row_index}: Error embedding image (ValueError): {ve}", exc_info=False)
                    img_embed_cell.value = f"画像形式エラー? ({ve})"
                except Exception as e:
                    logging.error(f"Row {row_index}: Unexpected error embedding image: {e}", exc_info=True)
                    img_embed_cell.value = "画像埋込エラー"

            else:
                # 画像ダウンロード/処理失敗
                logging.warning(f"Row {row_index}: Failed to download or prepare image from URL: {image_url}")
                img_embed_cell.value = "画像DL/処理失敗"
        else:
            # 画像URL取得失敗
            img_url_cell.value = error_message if error_message else "取得エラー"
            logging.error(f"Row {row_index}: Failed to get image URL. Error: {error_message}")
            # 画像URLが見つからなくてもD列には "-" が入っている

        # 1行処理完了表示
        current_elapsed_time = time.time() - overall_start_time
        print(f"\r処理完了: {processed_count}/{total_rows_with_urls} 件目 ({row_index}行目) - {url[:60]}... (経過: {current_elapsed_time:.1f} 秒)      ", flush=True)

        # 待機
        if sleep_interval > 0:
            logging.debug(f"Sleeping for {sleep_interval} seconds...")
            time.sleep(sleep_interval)

    print() # 最後の行の表示をクリアするための改行
    if processed_count == 0 and total_rows_with_urls > 0:
        print("有効なURLが見つかりましたが、処理は実行されませんでした（中断された可能性があります）。")
    elif processed_count > 0:
        total_elapsed_time = time.time() - overall_start_time
        print(f"\n全 {processed_count} 件のURL処理が完了しました。(合計時間: {total_elapsed_time:.2f} 秒)")

    return processed_count

def save_workbook(workbook: Workbook, file_path: str):
    """ワークブックを指定されたパスに保存する"""
    print(f"変更を '{file_path}' に保存しています...")
    t_save_start = time.time()
    try:
        workbook.save(file_path)
        print("保存が完了しました。")
        logging.info(f"Excel file saving completed in {time.time() - t_save_start:.3f}s")
    except PermissionError:
        print(f"\nエラー: ファイル '{file_path}' への書き込み権限がありません。")
        print("ファイルが開かれている場合は閉じてから再試行してください。")
        logging.error(f"Permission denied when saving Excel file: {file_path}")
    except Exception as save_e:
        print(f"\nエラー: ファイルの保存中に問題が発生しました: {save_e}")
        logging.error(f"Error saving Excel file '{file_path}': {save_e}", exc_info=True)

# ============================================
# 7. ロギング設定
# ============================================
def setup_logging(debug_mode: bool):
    """ロギングを設定する"""
    log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - [%(funcName)s:%(lineno)d] - %(message)s')
    logger = logging.getLogger()
    logger.handlers.clear() # 既存のハンドラをクリア

    # コンソールハンドラ
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(log_formatter)
    # デバッグモードならINFO、そうでなければWARNINGをコンソールに出力
    console_log_level = logging.INFO if debug_mode else logging.WARNING
    console_handler.setLevel(console_log_level)
    logger.addHandler(console_handler)

    # ファイルハンドラ (デバッグモード時のみ)
    if debug_mode:
        try:
            # 既存のログファイルがあれば削除
            if os.path.exists(DEBUG_LOG_FILE):
                os.remove(DEBUG_LOG_FILE)
            file_handler = logging.FileHandler(DEBUG_LOG_FILE, mode='w', encoding='utf-8')
            file_handler.setFormatter(log_formatter)
            file_handler.setLevel(logging.DEBUG) # ファイルにはDEBUGレベルまで記録
            logger.addHandler(file_handler)
            logger.setLevel(logging.DEBUG) # ロガー全体のレベルをDEBUGに設定
            print(f"デバッグログが有効です。ログファイル: '{DEBUG_LOG_FILE}'")
        except Exception as e:
            logging.error(f"デバッグログファイル '{DEBUG_LOG_FILE}' の準備に失敗しました: {e}")
            # ファイルハンドラが設定できなくてもコンソールには出力されるようにする
            logger.setLevel(console_log_level) # ロガーレベルをコンソールレベルに戻す
    else:
        logger.setLevel(logging.WARNING) # デバッグモードでない場合はWARNING以上のみ

    # ライブラリのログレベル調整 (requests, urllib3, selenium)
    logging.getLogger("requests").setLevel(logging.WARNING)
    logging.getLogger("urllib3").setLevel(logging.WARNING)
    logging.getLogger("selenium").setLevel(logging.WARNING) # WebDriver自体のログはWARNING以上

# ============================================
# 8. メイン実行関数
# ============================================
def main():
    """スクリプトのメイン実行関数"""
    overall_start_time = time.time()

    # --- 引数解析 ---
    parser = argparse.ArgumentParser(
        description='Excel内のURLから画像を取得し、指定列に埋め込みます。',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter # デフォルト値をヘルプに表示
    )
    parser.add_argument('input_file', help='処理対象のExcelファイルパス (.xlsx)')
    parser.add_argument('--sheet', default=0, help='処理対象のシート名またはインデックス (0始まり)')
    parser.add_argument('--width', type=int, default=DEFAULT_IMAGE_WIDTH,
                        help='埋め込む画像の幅 (px)')
    parser.add_argument('--sleep', type=float, default=POST_URL_SLEEP,
                        help='各URL処理後の待機時間 (秒)')
    parser.add_argument('--all', action='store_true', dest='process_all',
                        help='URL列が空になった時点で処理を中断せずに、ファイルの最後まで処理を試みる')
    parser.add_argument('--skip-selenium', action='store_true',
                        help='Selenium (WebDriver) を使用しない (requestsのみで処理)')
    parser.add_argument('--debug', action='store_true',
                        help=f'デバッグログを有効にし、{DEBUG_LOG_FILE} に出力する')
    args = parser.parse_args()

    # --- ロギング設定 ---
    setup_logging(args.debug)
    logging.info("Script execution started.")
    logging.info(f"Arguments: {args}")

    workbook: Optional[Workbook] = None
    sheet: Optional[Worksheet] = None
    processed_count = 0

    try:
        # --- Excelファイルの読み込みとシート選択 ---
        workbook, sheet = load_workbook_and_sheet(args.input_file, args.sheet)
        if not workbook or not sheet:
            return # エラーメッセージは load_workbook_and_sheet 内で表示済み

        # --- ヘッダー解析と列インデックス取得 ---
        required_headers = [URL_HEADER_NAME, IMAGE_URL_HEADER_NAME]
        col_indices = get_column_indices(sheet, required_headers)
        if not col_indices:
            return # エラーメッセージは get_column_indices 内で表示済み

        url_col_idx = col_indices[URL_HEADER_NAME]
        img_url_col_idx = col_indices[IMAGE_URL_HEADER_NAME]

        # --- 固定列の確認と表示 ---
        try:
            openpyxl.utils.column_index_from_string(IMAGE_EMBED_COL_LETTER)
            openpyxl.utils.column_index_from_string(HYPHEN_COL_LETTER)
            print(f"使用列: URL='{URL_HEADER_NAME}' ({get_column_letter(url_col_idx)}), "
                  f"ImgURL='{IMAGE_URL_HEADER_NAME}' ({get_column_letter(img_url_col_idx)}), "
                  f"ImgEmbed={IMAGE_EMBED_COL_LETTER}, Hyphen={HYPHEN_COL_LETTER}")
        except ValueError:
            print(f"エラー: 固定列名 ({IMAGE_EMBED_COL_LETTER}, {HYPHEN_COL_LETTER}) が無効です。")
            logging.error(f"Invalid fixed column letters: {IMAGE_EMBED_COL_LETTER}, {HYPHEN_COL_LETTER}")
            return

        # --- 既存結果のクリア ---
        clear_previous_results(sheet, url_col_idx, img_url_col_idx, IMAGE_EMBED_COL_LETTER, HYPHEN_COL_LETTER)

        # --- WebDriverの準備 (必要な場合) ---
        driver: Optional[webdriver.Chrome] = None
        if not args.skip_selenium:
            # WebDriverManagerを `with` 文で使い、初期化と終了を自動管理
            with WebDriverManager() as managed_driver:
                if managed_driver:
                    driver = managed_driver # 正常に初期化された場合
                    # --- Excel行処理の実行 ---
                    processed_count = process_excel_rows(
                        sheet, url_col_idx, img_url_col_idx,
                        IMAGE_EMBED_COL_LETTER, HYPHEN_COL_LETTER,
                        args.width, args.sleep, args.process_all, driver
                    )
                else:
                    # WebDriverの初期化に失敗した場合
                    print("警告: WebDriverの初期化に失敗したため、Seleniumを利用した処理はスキップされます。")
                    logging.warning("WebDriver initialization failed. Selenium-dependent operations will be skipped.")
                    # Seleniumなしで処理を試みる
                    processed_count = process_excel_rows(
                        sheet, url_col_idx, img_url_col_idx,
                        IMAGE_EMBED_COL_LETTER, HYPHEN_COL_LETTER,
                        args.width, args.sleep, args.process_all, None # driver=Noneを渡す
                    )
        else:
            # --skip-seleniumが指定された場合
            print("Seleniumの使用はスキップされました (--skip-selenium)。")
            logging.info("Selenium usage skipped due to --skip-selenium flag.")
            # --- Excel行処理の実行 (Seleniumなし) ---
            processed_count = process_excel_rows(
                sheet, url_col_idx, img_url_col_idx,
                IMAGE_EMBED_COL_LETTER, HYPHEN_COL_LETTER,
                args.width, args.sleep, args.process_all, None # driver=None
            )

        # --- ワークブックの保存 ---
        if processed_count > 0 or args.debug: # 処理件数が0でもデバッグモードなら保存試行
             if workbook:
                 save_workbook(workbook, args.input_file)
             else: # 通常発生しないはずだが念のため
                 logging.error("Workbook object is unexpectedly None before saving.")
        else:
             print("処理対象がなく、ファイルへの変更はありませんでした。保存はスキップします。")
             logging.info("No rows processed or changes made. Skipping file save.")


    except KeyboardInterrupt:
        print("\n処理がユーザーによって中断されました。")
        logging.warning("Process interrupted by user (KeyboardInterrupt).")
        # 中断時点までの変更を保存するかどうか選択させることも可能だが、
        # ここでは単純に終了する。必要なら保存処理を追加。
        # if workbook:
        #     save_workbook(workbook, args.input_file + "_interrupted.xlsx")

    except ImportError as ie:
         if 'selenium' in str(ie).lower():
             print("\nエラー: Seleniumライブラリがインストールされていません。")
             print("インストールするには、コマンドプロンプト/ターミナルで `pip install selenium` を実行してください。")
             print("また、Chromeブラウザと、それに対応するChromeDriverが必要です。")
         elif 'openpyxl' in str(ie).lower():
              print("\nエラー: openpyxlライブラリがインストールされていません。")
              print("インストールするには、 `pip install openpyxl` を実行してください。")
         elif 'pillow' in str(ie).lower():
              print("\nエラー: Pillowライブラリがインストールされていません。")
              print("インストールするには、 `pip install Pillow` を実行してください。")
         else:
             print(f"\nエラー: 必要なライブラリが見つかりません: {ie}")
         logging.error(f"Missing required library: {ie}", exc_info=True)

    except Exception as e:
        print(f"\n予期せぬエラーが発生しました: {e}")
        logging.exception("An unexpected error occurred in the main execution block.")

    finally:
        # --- ワークブックを閉じる ---
        if workbook:
            try:
                workbook.close()
                logging.info("Workbook closed.")
            except Exception as close_e:
                logging.error(f"Error closing workbook: {close_e}")

        overall_end_time = time.time()
        logging.info(f"Script execution finished. Total time: {overall_end_time - overall_start_time:.2f} seconds.")
        print(f"\nスクリプト全体の実行時間: {overall_end_time - overall_start_time:.2f} 秒")


if __name__ == "__main__":
    main()
